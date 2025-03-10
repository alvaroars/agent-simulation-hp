import os
import logging
import pandas as pd
import random
from itertools import permutations

try:
    from .simulator_dynamics import PhiFunction, VFunction, PhiGroup
except ImportError:
    from simulator_dynamics import PhiFunction, VFunction, PhiGroup
try:
    from .auxiliary import adjust_excel, generate_latex_table
except ImportError:
    from auxiliary import adjust_excel, generate_latex_table
# Additional imports to resolve missing modules and configuration variables
try:
    from openpyxl import load_workbook
except ImportError:
    from openpyxl import load_workbook

try:
    from openpyxl.drawing.image import Image
except ImportError:
    from openpyxl.drawing.image import Image

try:
    from .auxiliary import latex_settings, saving_routine_excel
except ImportError:
    from auxiliary import latex_settings, saving_routine_excel

try:
    from .config import tables_path, results_path, repo_path, figures_path
except ImportError:
    from config import tables_path, results_path, repo_path, figures_path

import numpy as np
import matplotlib.pyplot as plt

# Latex column dictionary
LATEX_COLUMN_DICT = {
    "Expected value random": r"$\hat{\mathcal{A}}(\Phi_R)$",
    "Expected iterations random": r"$\hat{\tau}^R$",
    "Expected value best": r"$\hat{\mathcal{A}}(\Phi_B)$",
    "Expected iterations best": r"$\hat{\tau}^B$",
    "Best outperform random": r"$\hat{\mathcal{O}}_{\Phi_B, \Phi_R}$",
    "Distinct in best heuristic (sorted)": r"$N_B$",
    "Intersection": r"$N_I$",
    "Mean ability best": r"$$\widehat{\textnormal{mean}}_{\Phi_B}\left(\mathcal{A}(\phi)\right)$$",
    "Min ability best": r"$\widehat{\textnormal{min}}_{\Phi_B}\left(\mathcal{A}(\phi)\right)$",
    "Max ability best": r"$\widehat{\textnormal{max}}_{\Phi_B}\left(\mathcal{A}(\phi)\right)$",
    "Mean ability random": r"$\widehat{\textnormal{mean}}_{\Phi_R}\left(\mathcal{A}(\phi)\right)$",
    "Min ability random": r"$\widehat{\textnormal{min}}_{\Phi_R}\left(\mathcal{A}(\phi)\right)$",
    "Max ability random": r"$\widehat{\textnormal{max}}_{\Phi_R}\left(\mathcal{A}(\phi)\right)$"
}


def _create_excel_from_df(df, file_name, open_excel=False) -> None:
    """Create an Excel file from a DataFrame.


    :param df: The DataFrame to be converted to Excel
    :type df: pandas.DataFrame
    :param file_name: The name of the file to be created
    :type file_name: str
    :param open_excel: Whether to open the Excel file after creation
    :type open_excel: bool
    """
    # Create a deep copy of the DataFrame, so the new rows do not affect the original one
    df_excel = df.copy(deep=True)
    # Add final row with the average of the columns
    df_excel.loc['Average'] = df.mean()
    df_excel.loc['Std'] = df.std()
    # Write results to file
    try:
        df_excel.to_excel(file_name)
    except PermissionError:
        input(
            f"Permission denied: Unable to save the file at {file_name}. Try closing the file and press Enter to continue...")
        df_excel.to_excel(file_name)
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        raise
    adjust_excel(file_name, header="large", filter="all", align='center', open_excel=open_excel)


# Testing functions
# noinspection PyPep8Naming
def test_individual_h(h, V, prints=False) -> tuple[float, float]:
    """Test an individual heuristic.

    :param h: The heuristic to test.
    :type h: list
    :param V: The :math:`V` function.
    :type V: VFunction
    :param prints: Whether to print the results.
    :type prints: bool

    :return: The expected value of the heuristic.
    :rtype: tuple[float, float]
    """
    phi = PhiFunction(V=V, heuristics=h)
    if prints:
        print("Heuristic: ", h)
        print("Expected value: ", phi.expected_value)
        print("")
    else:
        return phi.expected_value


# Testing group dynamics
def test_group_h(heuristics, V, no_random=10, no_elements=None, additional_best=None, analysis=False,
                 less_ability_coef=None, no_less=0, alternative_stop=False, stop=None, delta_rho=None,
                 analysis_ability_difference=False) -> tuple:
    """Test the group dynamics of a set of heuristics.

    :param heuristics: The heuristics to test.
    :type heuristics: list
    :param V: The :math:`V` function.
    :type V: VFunction
    :param no_random: The number of random heuristics to test.
    :type no_random: int
    :param no_elements: The number of elements to test.
    :type no_elements: int
    :param additional_best: The additional best heuristics to test.
    :type additional_best: list
    :param analysis: Whether to analyze the results.
    :type analysis: bool
    :param less_ability_coef: The less ability coefficient.
    :type less_ability_coef: float
    :param no_less: The number of less ability elements.
    :type no_less: int
    :param alternative_stop: Whether to use an alternative stop condition.
    :type alternative_stop: bool
    :param stop: The stop condition.
    :type stop: int
    :param delta_rho: List of delta ratios for the group dynamics.
    :type delta_rho: list
    :param analysis_ability_difference: Whether to analyze the ability difference.
    :type analysis_ability_difference: bool

    :return: The expected value of the heuristics.
    :rtype: tuple
    """

    n = V.n
    # l is the maximum number that each element in the heuristic can take
    l = max(max(heuristics))
    # k is the number of elements in the heuristic
    k = len(heuristics[0])

    if no_elements is None:
        no_elements = no_random
    random_pool_h = heuristics
    best_pool_phi = [PhiFunction(V=V, heuristics=h) for h in random_pool_h]
    random_heur = random.sample(random_pool_h, no_random)
    phi_random = PhiGroup._create_random_phi(V, random_heur, less_ability_coef=less_ability_coef,
                                             no_less=no_less, delta_rho=delta_rho)
    # Additional best phi functions, which are not included in the random pool, although highly unlikely if they were
    if additional_best is not None:
        best_pool_phi += additional_best

    # Best phis
    best_phi, best_heuristics, intersection, distinct_heur_sorted, distinct_heur_exp = (
        V.best_phis(heuristics=heuristics, no_elements=no_elements, additional_best=additional_best,
                    delta_rho=delta_rho))
    random_group = PhiGroup(phi_list=phi_random, is_best=False)
    best_group = PhiGroup(phi_list=best_phi, is_best=True)
    ev_random = random_group.expected_value_group(alternative_stop=alternative_stop, stop=stop)
    ev_best = best_group.expected_value_group(alternative_stop=alternative_stop, stop=stop)
    if analysis_ability_difference:
        ab_stats_best = best_group.analyze_ability_difference()
        ab_stats_random = random_group.analyze_ability_difference()
    if analysis:
        additional_best_no = len(additional_best) if additional_best is not None else 0
        if additional_best_no > 0 and additional_best[0].region is not None:
            search = len(additional_best[0].region)
        else:
            search = 5  # Default search
        name = f"phi_comparison_trial_V_random"
        name += f"_N_{no_random}_n_{n}_l_{l}_addPhi_{additional_best_no}_search_{search}_"
        name += f"less_{less_ability_coef}_noless_{no_less}_alternative_{alternative_stop}_"
        name += f"stop_{stop}.pdf"
        best_group.plot_comparison(random_group, title=r"$\Phi_B$ vs $\Phi_R$",
                                   name=name)
        V.plot(name="plot_V_" + name)
        print("Random group vs Best group")
        if delta_rho is None:
            for i in range(n):
                print(
                    f"{i} -> {(random_group(i)[0], f'{V(random_group(i)[0]):.3f}')} vs {(best_group(i)[0], f'{V(best_group(i)[0]):.3f}')}")
        best_message = "Best heuristics: " + str(best_heuristics)
        if intersection is not None:
            best_message += f" and {intersection} additional best phi functions"
        print("Distinct in best heuristics: ", distinct_heur_sorted, " Distinct in expected values: ",
              distinct_heur_exp)
        print("Random heuristics: ", random_heur)
        # Extract the different numbers in the random heuristics and the missing ones
        random_heur_numbers = set([i for h in random_heur for i in h])
        missing_heur = [i for i in range(1, l + 1) if i not in random_heur_numbers]
        print("Random heuristics unique: ", random_heur_numbers)
        print("Missing heuristics: ", missing_heur)
        # The same for the best heuristics
        best_heur_numbers = set([i for h in best_heuristics for i in h])
        missing_best_heur = [i for i in range(1, l + 1) if i not in best_heur_numbers]
        print("Best heuristics unique: ", best_heur_numbers)
        print("Missing best heuristics: ", missing_best_heur)
        print(best_message)
        print("Random expected value: ", ev_random)
        print("Random expected value, individual: ", [phi.expected_value[0] for phi in phi_random])
        print("Best expected value: ", ev_best)
        print("Best expected value, individual: ", [phi.expected_value[0] for phi in best_phi])
        print("Best heuristics expected value, individual: ",
              [PhiFunction(heuristics=h, V=V).expected_value[0] for h in best_heuristics])
        print("")
        # input("Press Enter to continue...")
    if analysis_ability_difference:
        return ev_random, ev_best, intersection, distinct_heur_sorted, distinct_heur_exp, ab_stats_best, ab_stats_random
    else:
        return ev_random, ev_best, intersection, distinct_heur_sorted, distinct_heur_exp


def _generate_caption(file_name_xlsx, no_Vs, no_random, n, l, percentage_random, additional_best_no,
                      less_ability_coef, no_less, alternative_stop, stop, delta_rho, params=None):
    """Generate a caption for the results of the expected values and the number of iterations.

    :param file_name_xlsx: The name of the file to be created.
    :type file_name_xlsx: str
    :param no_Vs: The number of :math:`V` functions to test.
    :type no_Vs: int
    :param no_random: The number of random heuristics to test.
    :type no_random: int
    :param n: The number of elements in the function :math:`V`.
    :type n: int
    :param l: The maximum number that each element in the heuristic can take.
    :type l: int
    :param percentage_random: The percentage of random :math:`V` functions.
    :type percentage_random: float
    :param additional_best_no: The number of additional best heuristics to test.
    :type additional_best_no: int
    :param less_ability_coef: The less ability coefficient.
    :type less_ability_coef: float
    :param no_less: The number of less ability elements.
    :type no_less: int
    :param alternative_stop: Whether to use an alternative stop condition.
    :type alternative_stop: bool
    :param stop: The stop condition.
    :type stop: float
    :param delta_rho: The delta-rho ratio.
    :type delta_rho: list   

    :return: The caption.
    :rtype: str
    """

    # Readable caption
    if percentage_random == 1:
        phi_str = ""
    elif percentage_random == 0 and params is None:
        phi_str = ", $V$ functions are parametrically generated"
    elif params is not None:
        phi_str = r", $V$ function is parametrically generated with $\alpha, \beta, \gamma, x_0, \delta = " + \
                  f"({params[0]}, {params[1]}, {params[2]}, {params[3]}, {params[4]})$"
    else:
        phi_str = f", {percentage_random * 100}\\% percentage of random $\\\phi$ functions"
    add_best_str = f", {additional_best_no} additional $\\phi$ functions" if additional_best_no > 0 else ""
    less_ability_str = f", {less_ability_coef} less ability coefficient, {no_less} number of less ability elements" if less_ability_coef is not None else ""
    alternative_stop_str = ", stop when a disagreement cycle is detected" if alternative_stop else ""
    stop_str = f", stop set to {stop}" if stop is not None else ""
    delta_rho_str = f", delta-rho ratios $A_R$={delta_rho[0]} and $A_B$={delta_rho[1]}" if delta_rho is not None else ""

    caption = f"Results of the expected values and the number of iterations with $n$={n}, $l$={l}, " + \
              f"$M$={no_Vs} number of trials, " + \
              f"$N_1$={no_random} elements in the random group" + \
              phi_str + \
              add_best_str + \
              less_ability_str + \
              alternative_stop_str + \
              stop_str + \
              delta_rho_str + \
              ". " + \
              "In parenthesis, the standard deviation of the expected values."

    return caption


def _create_latex_table(df, file_name_xlsx, caption, all=False, precision=3):
    """
    Create a :math:`\\LaTeX` table from a data frame.

    :param df: The data frame to be converted to a :math:`\\LaTeX` table.
    :type df: pandas.DataFrame
    :param file_name_xlsx: The name of the file to be created.
    :type file_name_xlsx: str
    :param caption: The caption of the table.
    :type caption: str
    :param all: Whether to show all columns.
    :type all: bool
    :param precision: The precision of the table.
    :type precision: int

    :return: None
    :rtype: None
    """
    columns_latex = [LATEX_COLUMN_DICT[col] for col in df.columns if col in LATEX_COLUMN_DICT]
    columns_old = [col for col in df.columns if col in LATEX_COLUMN_DICT]
    df_latex = pd.DataFrame(columns=columns_latex, index=df.index, data=df[columns_old].values)
    if not all:
        # Keep only the first 5 columns
        df_latex = df_latex.iloc[:, :5]
    df_formatted = df_latex.copy(deep=True)
    df_formatted = df_formatted.astype(str)

    for i in range(len(df_latex.columns)):
        avg = df_latex.iloc[:, i].mean()
        std = df_latex.iloc[:, i].std()
        p = precision
        # Integers with no decimal points, else 4 decimal points
        if avg == int(avg):
            avg = int(avg)
            df_formatted.loc['Average', df_formatted.columns[i]] = f"{avg:.{p}f} ({std:.{p}f})"
        else:
            df_formatted.loc['Average', df_formatted.columns[i]] = f"{avg:.{p}f} ({std:.{p}f})"
    # Keep only the average row
    df_formatted = df_formatted.loc[['Average']]
    output_path = os.path.join(tables_path, file_name_xlsx.replace(".xlsx", ""))
    label = "tab:" + file_name_xlsx.replace(".xlsx", "")
    generate_latex_table(df_formatted, output_path=output_path, caption=caption, label=label)


# Results functions
def excel_results(n, l, k, no_Vs=100, latex=True, percentage_random=1, no_random=10,
                  add_radial_search=False, additional_best_no=0, params=None, search=5,
                  less_ability_coef=None, no_less=0, alternative_stop=False, stop=None,
                  delta_rho=None, analysis=False, function=None, analysis_ability_difference=False,
                  open_excel=True) -> None:
    """Create an Excel file with the results of the expected values and the number of iterations for each phi function.

    :param no_Vs: Number of :math:`V` functions to test.
    :type no_Vs: int
    :param latex: Flag to create a :LaTeX: table with the results.
    :type latex: bool
    :param percentage_random: Percentage of random :math:`V` functions.
    :type percentage_random: float
    :param add_radial_search: Flag to add a radial search to the pool of best phis.
    :type add_radial_search: bool
    :param additional_best_no: Number of additional best phi functions to add to the pool of best phis.
    :type additional_best_no: int
    :param params: Parameters for the deterministic V function. If entered, the random V functions are ignored, it
     returns the results with just one function V.
    :param search: Number of elements in the region for the additional phis.
    :type search: int
    :param less_ability_coef: The less ability coefficient. See :py:class:`.simulator_dynamics.PhiFunction`.
    :type less_ability_coef: float
    :param no_less: The number of less ability elements.
    :type no_less: int
    :param alternative_stop: Flag to use the alternative stop condition (stop when a disagreement cycle is detected).
    :type alternative_stop: bool
    :param stop: The stop condition.
    :type stop: int
    :param analysis_ability_difference: Flag to analyze the ability difference.
    :type analysis_ability_difference: bool
    :param delta_rho: List of deltas and rhos for the group dynamics. In particular, if provided, it contains:
        - delta_rho[0]: The ratio of deltas and rhos for the random group.
        - delta_rho[1]: The ratio of deltas and rhos for the best group.
    :type delta_rho: list 
    :param analysis: Flag to analyze the phi functions.
    :type analysis: bool
    :param function: The function :math:`V`.
    :type function: function
    :param open_excel: Flag to open the Excel file.
    :type open_excel: bool

    :return: None
    """
    if delta_rho is not None and not add_radial_search:
        raise ValueError("The delta_rho parameter is only available when add_radial_search is True")
    if params is not None or function is not None:
        no_Vs = 1
        percentage_random = 0
        print(Warning("The parameter no_Vs, percentage_random are ignored when params is provided"))

    results = dict()
    if analysis_ability_difference:
        results_ab_diff = dict()
    for i in range(no_Vs):
        print(f"V function {i}")
        if i / no_Vs < percentage_random:
            V = VFunction(n=n, random=True)
        elif params is not None:
            V = VFunction(n=n, params=params)
        else:
            V = VFunction(n=n, function=function)

        index = V.return_index(case=i)
        # Store the picture of V, one for each V
        if i / no_Vs >= percentage_random:
            V.plot(name=f"V_{i}.png", saving_folder=repo_path)

        # Define best agents
        phi_best = V.create_additional_bests(add_radial_search=add_radial_search,
                                             additional_best_no=additional_best_no,
                                             search=search, delta_rho=delta_rho)

        # Compare the expected values of the best heuristics with random heuristics
        if delta_rho is None:
            pool_tools = PhiGroup.create_pool_tools(l=l, k=k)
        else:
            length = 0
            repeat = 0
            while length < no_random:
                repeat += 1
                pool_tools = PhiGroup.create_pool_tools(delta_rho_ratio=delta_rho[0], repeat=repeat)
                length = len(pool_tools)
            if repeat > 1:
                print(f"Repeat {repeat} times to get {no_random} heuristics")
        if analysis_ability_difference:
            ev_random, ev_best, intersection, distinct_heur_sorted, distinct_heur_exp, ab_stats_best, ab_stats_random = (
                test_group_h(pool_tools, V, no_random=no_random, additional_best=phi_best,
                             less_ability_coef=less_ability_coef, no_less=no_less, alternative_stop=alternative_stop,
                             stop=stop, delta_rho=delta_rho, analysis=analysis,
                             analysis_ability_difference=analysis_ability_difference))
            results_ab_diff[index] = (*ab_stats_best, *ab_stats_random)
        else:
            ev_random, ev_best, intersection, distinct_heur_sorted, distinct_heur_exp = (
                test_group_h(pool_tools, V, no_random=no_random, additional_best=phi_best,
                             less_ability_coef=less_ability_coef, no_less=no_less, alternative_stop=alternative_stop,
                             stop=stop, delta_rho=delta_rho, analysis=analysis))

        best_beat_heur = ev_best[0] > ev_random[0]

        # Store results in a dictionary
        results[index] = (ev_random[0], ev_random[1], ev_best[0], ev_best[1], best_beat_heur, distinct_heur_sorted,
                          distinct_heur_exp, intersection)

    df = pd.DataFrame.from_dict(results, orient='index', columns=['Expected value random',
                                                                  'Expected iterations random',
                                                                  'Expected value best',
                                                                  'Expected iterations best',
                                                                  'Best outperform random',
                                                                  'Distinct in best heuristic (sorted)',
                                                                  'Distinct in best heuristic (expected)',
                                                                  'Intersection'])

    if delta_rho is None:
        file_name = (f'results_M_{no_Vs}_randomVperc_{percentage_random}_N_{no_random}_n_{n}_l_{l}_'
                     f'addPhi_{additional_best_no}_search_{search}_less_{less_ability_coef}_noless_{no_less}_'
                     f'alternative_{alternative_stop}_stop_{stop}.xlsx')
    else:
        file_name = (f'results_M_{no_Vs}_randomVperc_{percentage_random}_N_{no_random}_n_{n}_l_{l}_'
                     f'delta_rho_{delta_rho}.xlsx')
    _create_excel_from_df(df, os.path.join(results_path, file_name))

    if analysis_ability_difference:
        df_ab_diff = pd.DataFrame.from_dict(results_ab_diff, orient='index',
                                            columns=['Mean ability best',
                                                     'Min ability best',
                                                     'Max ability best',
                                                     'Mean ability random',
                                                     'Min ability random',
                                                     'Max ability random'])
        file_name_ab_diff = file_name.replace(".xlsx", "_ab_diff.xlsx")
        _create_excel_from_df(df_ab_diff, os.path.join(results_path, file_name_ab_diff))

    # Add the pictures to the Excel file so it fits inside a cell on the right
    wb = load_workbook(os.path.join(results_path, file_name))
    ws = wb.active
    for i in range(no_Vs):
        if i / no_Vs >= percentage_random:
            # Adjusted image size
            img = Image(f"V_{i}.png")
            img.width = 100
            img.height = 100
            # Count las column
            no_columns = len(df.columns) + 1
            col = chr(65 + no_columns)  # That goes through A, B, ...
            ws.add_image(img, f'{col}{i + 2}')
            ws.row_dimensions[i + 2].height = 100  # adjust cell size to fit the image
            ws.column_dimensions[col].width = 20

    excel_path = os.path.join(results_path, file_name)
    saving_routine_excel(wb, excel_path, open_excel=open_excel)

    if latex:
        if latex == 'all':
            all = True
            precision = 2
        else:
            all = False
            precision = 3
        caption = _generate_caption(file_name_xlsx=file_name, no_Vs=no_Vs, no_random=no_random, n=n,
                                    l=l, percentage_random=percentage_random,
                                    additional_best_no=additional_best_no,
                                    less_ability_coef=less_ability_coef, no_less=no_less,
                                    alternative_stop=alternative_stop, stop=stop, params=params,
                                    delta_rho=delta_rho)
        _create_latex_table(df, file_name_xlsx=file_name, caption=caption, all=all, precision=precision)
        if analysis_ability_difference:
            caption_ab_diff = caption.replace("Results of the expected values and the number of iterations with",
                                              "Statistics of the ability difference between the best and random heuristics with")
            file_name_ab_diff = file_name.replace(".xlsx", "_ab_diff.xlsx")
            _create_latex_table(df_ab_diff, file_name_xlsx=file_name_ab_diff, caption=caption_ab_diff,
                                all=True, precision=2)


def test_V(params=None, trials=100, no_random=10, n=200, l=12, k=3, add_radial_search=False,
           additional_best_no=0, search=5, less_ability_coef=None, no_less=0,
           alternative_stop=False, stop=None, latex=True, analysis=False, delta_rho=None,
           open_excel=True) -> None:
    """Test a fixed V taking several trials for the random group.

    :param params: Parameters for the deterministic :math:`V` function. If entered, the random :math:`V` functions are ignored, it
     returns the results with just one function :math:`V`.
    :type params: tuple
    :param trials: Number of trials for the random group.
    :type trials: int
    :param no_random: Number of elements in the random group.
    :type no_random: int
    :param n: Number of elements in the function :math:`V`.
    :type n: int
    :param l: Maximum number that each element in the heuristic can take.
    :type l: int
    :param add_radial_search: Flag to add a radial search to the pool of best phis.
    :type add_radial_search: bool
    :param additional_best_no: Number of additional best phi functions to add to the pool of best phis.
    :type additional_best_no: int
    :param search: Number of elements in the region for the additional phis.
    :type search: int
    :param less_ability_coef: Less ability coefficient.
    :type less_ability_coef: float
    :param no_less: Number of less ability elements.
    :type no_less: int
    :param alternative_stop: Flag to use the alternative stop condition (stop when a disagreement cycle is detected).
    :type alternative_stop: bool
    :param stop: Stop condition.
    :type stop: int
    :param latex: Flag to create a :math:`\LaTeX` table with the results.
    :type latex: bool
    :param analysis: Flag to analyze the phi functions.
    :type analysis: bool
    :param delta_rho: List of deltas and rhos for the group dynamics.
    :type delta_rho: list

    """
    V = VFunction(n=n, params=params)
    if delta_rho is None:
        pool_tools = PhiGroup.create_pool_tools(l=l, k=k)
    else:
        pool_tools = PhiGroup.create_pool_tools(delta_rho_ratio=delta_rho[0], repeat=max(1, int(100 / no_random)))
    best_phi, best_heuristics, intersection, distinct_heur_sorted, distinct_heur_exp = (
        V.best_phis(heuristics=pool_tools, no_elements=no_random, additional_best=None, delta_rho=delta_rho))
    best_group = PhiGroup(phi_list=best_phi, is_best=True)
    if analysis:
        if additional_best_no == 0 and delta_rho is None:
            print(f"Best heuristics: {best_heuristics}")
            ev_best = best_group.expected_value_group(alternative_stop=alternative_stop, stop=stop)
            print("Best expected value: ", ev_best)
            print("Best expected value, individual: ", [phi.expected_value[0] for phi in best_phi])
            print("Best heuristics expected value, individual: ",
                  [PhiFunction(heuristics=h, V=V).expected_value[0] for h in best_heuristics])
            print("")
        elif delta_rho is not None:
            print(f"Best delta_rho: {delta_rho}")
    results = dict()
    for i in range(trials):
        print(f"Trial {i}")
        add_best = V.create_additional_bests(add_radial_search=add_radial_search, additional_best_no=additional_best_no,
                                             search=search, delta_rho=delta_rho)
        # Note that this is random too
        if analysis and additional_best_no > 0:  # Analyze the phi functions
            add_best_group = PhiGroup(phi_list=add_best)
            add_best_group.analyze_phi_region()

        best_phi, best_heuristics, intersection, distinct_heur_sorted, distinct_heur_exp = (
            V.best_phis(heuristics=pool_tools, no_elements=no_random, additional_best=add_best,
                        delta_rho=delta_rho, best_heuristics=best_heuristics))
        best_group = PhiGroup(phi_list=best_phi, is_best=True)
        ev_best = best_group.expected_value_group(alternative_stop=alternative_stop, stop=stop)

        random_sample = random.sample(pool_tools, no_random)
        phi_random = PhiGroup._create_random_phi(V, random_sample, less_ability_coef=less_ability_coef, no_less=no_less)
        random_group = PhiGroup(phi_list=phi_random, is_best=False)
        ev_random = random_group.expected_value_group(alternative_stop=alternative_stop, stop=stop)
        if analysis:
            best_group.plot_comparison(random_group, alternative_stop=alternative_stop, stop=stop,
                                       title=r"$\Phi_B$ vs $\Phi_R$",
                                       name=f"phi_comparison_trial_{i}_V_{params}_N_{no_random}_n_{n}_l_{l}_addPhi_{additional_best_no}_"
                                            f"search_{search}_less_{less_ability_coef}_noless_{no_less}_alternative_{alternative_stop}_"
                                            f"stop_{stop}.pdf", sizes=(17, 18, 17, 17, 16, (10, 6)))
            for i in range(n):
                print(
                    f"{i} -> {(random_group(i)[0], f'{V(random_group(i)[0]):.3f}')} vs {(best_group(i)[0], f'{V(best_group(i)[0]):.3f}')}")
            best_message = "Best heuristics: " + str(best_heuristics)
            if intersection is not None:
                best_message += f" and {intersection} additional best phi functions"
            print("Distinct in best heuristics: ", distinct_heur_sorted, " Distinct in expected values: ",
                  distinct_heur_exp)
            random_heur = random_sample
            print("Random heuristics: ", random_heur)
            # Extract the different numbers in the random heuristics and the missing ones
            random_heur_numbers = set([i for h in random_heur for i in h])
            missing_heur = [i for i in range(1, l + 1) if i not in random_heur_numbers]
            print("Random heuristics unique: ", random_heur_numbers)
            print("Missing heuristics: ", missing_heur)
            # The same for the best heuristics
            best_heur_numbers = set([i for h in best_heuristics for i in h])
            missing_best_heur = [i for i in range(1, l + 1) if i not in best_heur_numbers]
            print("Best heuristics unique: ", best_heur_numbers)
            print("Missing best heuristics: ", missing_best_heur)
            print(best_message)
            print("Random expected value: ", ev_random)
            print("Random expected value, individual: ", [phi.expected_value[0] for phi in phi_random])

        if analysis and additional_best_no == 0:
            print([phi.heuristics for phi in phi_random])
        best_beat_heur = ev_best[0] > ev_random[0]

        # Store results in a dictionary
        results[i] = (ev_random[0], ev_random[1], ev_best[0], ev_best[1], best_beat_heur, distinct_heur_sorted,
                      distinct_heur_exp, intersection)

    df = pd.DataFrame.from_dict(results, orient='index', columns=["Expected value random", "Expected iterations random",
                                                                  'Expected value best', 'Expected iterations best',
                                                                  'Best outperform random',
                                                                  'Distinct in best heuristic (sorted)',
                                                                  'Distinct in best heuristic (expected)',
                                                                  'Intersection'])
    if delta_rho is None:
        file_name = f"test_V_{params}_M_{trials}_N_{no_random}_n_{n}_l_{l}_addPhi_{additional_best_no}_" \
                    f"search_{search}_less_{less_ability_coef}_noless_{no_less}_" \
                    f"alternative_{alternative_stop}_stop_{stop}.xlsx"
    else:
        file_name = f"test_V_{params}_M_{trials}_N_{no_random}_n_{n}_l_{l}_delta_rho_{delta_rho}.xlsx"
    _create_excel_from_df(df, os.path.join(results_path, file_name), open_excel=False)
    # Add the image of the function V inside the Excel file
    V.plot(name=f"V_{params}.png")
    wb = load_workbook(os.path.join(results_path, file_name))
    ws = wb.active
    try:
        img = Image(f"V_{params}.png")
    except FileNotFoundError:
        print(f"File V_{params}.png not found, generating it")
        V.plot(name=f"V_{params}.png", saving_folder=figures_path)
    img = Image(os.path.join(figures_path, f"V_{params}.png"))
    img.width = 100
    img.height = 100
    ws.add_image(img, 'A1')
    wb.save(os.path.join(results_path, file_name))
    # Open excel
    if open_excel:
        os.system(f'start excel "{os.path.join(results_path, file_name)}"')
    if latex:
        if latex == 'all':
            all = True
        else:
            all = False
        caption = _generate_caption(file_name_xlsx=file_name, no_Vs=trials, no_random=no_random, n=n,
                                    l=l, percentage_random=0,
                                    additional_best_no=additional_best_no,
                                    less_ability_coef=less_ability_coef, no_less=no_less,
                                    alternative_stop=alternative_stop, stop=stop, params=params,
                                    delta_rho=delta_rho)
        _create_latex_table(df, file_name_xlsx=file_name, caption=caption, all=all)


def plots_convergence_from_excel(file_name, column_names, title=None, saving_path=None,
                                 saving_name=None, ma_window=50, ylim=None, sizes=None,
                                 group_statistics=False):
    """
    Compute the cumulative mean as a function of the number of V functions from an Excel file.
    
    :param file_name: The name of the file to be created.
    :type file_name: str
    :param column_names: The names of the columns to be plotted.
    :type column_names: list
    :param title: The title of the plot.
    :type title: str
    :param saving_path: The path to save the plot.
    :type saving_path: str
    :param saving_name: The name of the file to be saved.
    :type saving_name: str
    :param ma_window: The window size for the moving average.
    :type ma_window: int
    :param ylim: The y-axis limits.
    :type ylim: tuple
    :param sizes: The sizes of the plot.
    :type sizes: tuple
    :param group_statistics: Whether to plot the group statistics.
    :type group_statistics: bool

    :return: None
    :rtype: None
    """
    df = pd.read_excel(file_name, index_col=0)
    if group_statistics:
        file_name_stats = file_name.replace(".xlsx", "_ab_diff.xlsx")
        df_stats = pd.read_excel(file_name_stats, index_col=0)
        df_total = pd.concat([df, df_stats], axis=1)
    else:
        df_total = df
    values = df_total[column_names].values
    # Exclude rows corresponding to the average and standard deviation (last two rows)
    values = values[:-2]
    ys_mean = np.cumsum(values, axis=0) / np.arange(1, len(values) + 1)[:, None]
    if len(values) < ma_window:
        raise ValueError(f"Not enough data points ({len(values)}) for the moving average with window size {ma_window}.")
    df_values = pd.DataFrame(values)
    ys_mav = df_values.rolling(window=ma_window, min_periods=1).mean().values
    if sizes is not None:
        labelsize = sizes[0]
        title_size = sizes[1]
        ysize = sizes[2]
        xsize = sizes[3]
        legend_size = sizes[4]
        figsize = sizes[5]
    else:
        labelsize = 12
        title_size = 13
        ysize = 12
        xsize = 12
        legend_size = 12
        figsize = (10, 6)
    latex_settings(plt, labelsize=labelsize, figsize=figsize)
    for i, column_name in enumerate(column_names):
        if column_name in ["Expected value random", 'Expected value best']:
            if column_name.endswith("random"):
                phi = r"\Phi_R"
                color = "blue"
            else:
                phi = r"\Phi_B"
                color = "red"
            if not group_statistics:
                label_mav = r"$\mathbb{E}_\nu(V\circ{\phi^{" + phi + r"}})^\textsc{ma}$"
                plt.plot(ys_mav[:, i], color=color, linestyle="-", label=label_mav, linewidth=1)
            label_mean = r"$\hat{\mathcal{A}}_M(" + phi + r")$"
            plt.plot(ys_mean[:, i], color=color, linestyle="--", label=label_mean, linewidth=2)
        elif column_name == "Best outperform random":
            label_mean = r"$\hat{\mathcal{O}}_{\Phi_B, \Phi_R}$"
            color = 'black'
            plt.plot(ys_mean[:, i], color=color, linestyle="--", label=label_mean, linewidth=2)
        elif column_name in ["Mean ability best", "Min ability best", "Max ability best", "Mean ability random",
                             "Min ability random", "Max ability random"]:
            if column_name.endswith("best"):
                phi = r"{\Phi_B}"
                color = "red"
            else:
                phi = r"{\Phi_R}"
                color = "blue"
            if column_name.startswith("Mean"):
                operator = r"$\widehat{\textnormal{mean}}_"
                style = "-"
                width = 2
            elif column_name.startswith("Min"):
                operator = r"$\widehat{\textnormal{min}}_"
                style = ":"
                width = 1
            elif column_name.startswith("Max"):
                operator = r"$\widehat{\textnormal{max}}_"
                style = "--"
                width = 1
            label_mean = operator + phi + r"\left(\mathcal{A}(\phi)\right)$"
            plt.plot(ys_mean[:, i], color=color, linestyle=style, label=label_mean, linewidth=width)
        else:
            raise NotImplementedError(f"Column name {column_names} not implemented")
    if ylim is not None:
        plt.ylim(ylim)
    if group_statistics:
        lengend_pos = None
        legend_size = 10
    else:
        lengend_pos = None

    plt.legend(fontsize=legend_size, loc=lengend_pos)
    plt.ylabel("Ability estimate", fontsize=ysize)
    plt.xlabel(r"$M$", fontsize=xsize)
    split_name = file_name.split("\\")[-1]
    if title is None:
        title = split_name
    plt.title(title, fontsize=title_size)
    if saving_path is not None:
        if saving_name is None:
            title = title.replace(" ", "_")
            saving_name = title + ".pdf"
        else:
            if saving_name.endswith(".pdf"):
                saving_name = saving_name
            else:
                saving_name = saving_name + ".pdf"
        plt.savefig(os.path.join(saving_path, saving_name))
    else:
        plt.show()
    plt.tight_layout()
    plt.close()
