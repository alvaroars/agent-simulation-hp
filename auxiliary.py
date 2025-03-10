import logging
import os
import subprocess
import warnings

from openpyxl import load_workbook
from openpyxl.styles import Alignment


def latex_settings(plt, mathpazo=True, labelsize=12, figsize=(10, 6)):
    """Set the LaTeX settings for the plots."""
    plt.figure(figsize=figsize, tight_layout=True)
    plt.rcParams.update({
        'text.usetex': True,
        'font.family': 'serif',
        'font.serif': ['Palatino'] if mathpazo else ['Computer Modern Roman'],
        'xtick.labelsize': labelsize,
        'ytick.labelsize': labelsize,
    })


def adjust_excel(excel_path, header=None, filter="first", align='first_left', wrap_column=False,
                 open_excel=False):
    """Adjust the column widths of an Excel file or worksheet and align all cells. It also adds a filter to the columns if required.

    :param excel_path: Path to the Excel file.
    :type excel_path: str
    :param header: Whether to treat the header, first row, as a normal row or not. If 'adjust', the width of the columns
        is adjusted to the length of the header. If 'exclude', the header is excluded from the adjustment. If 'large',
        the header is treated as a normal row, but the text is wrapped. This is useful for large headers. If None, the
        header is treated as a normal row.
    :type header: str or None
    :param filter: Whether to add a filter to the columns. It can be `all`, `first`, `first_two`, `last`, or `none`, optional (default
        is `first`).
    :type filter: str
    :param align: Alignment of the cells. It can be `first_left`, `center`, or `left`, optional
     (default is `first_left`). If `first_left`, the first column is left aligned, the rest are centered.
    :type align: str
    :param wrap_column: Whether to wrap the text in the header, optional (default is `False`). If `True`, the text in the
        header is wrapped.
    :type wrap_column: bool
    """
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    # Adjust the width of the columns
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        # First cell, we check if its empty
        if column[0].value is not None and len(str(column[0].value)) > 1 and header == "adjust":
            max_length = len(column[0].value)
        elif column[0].value is not None and len(str(column[0].value)) > 1 and header == "large":
            max_length = len(column[0].value) // 2
        else:
            for cell in column:
                if header == "exclude" and cell.row == 1:
                    continue
                # For float, take into account the format
                if isinstance(cell.value, float):
                    # Try to mimic general Excel format to get the number of characters
                    if abs(cell.value) < 1e-4:  # Scientific notation
                        # String representation of the number in scientific notation
                        string_sci = f"{cell.value:.5e}"
                        cell_length = len(string_sci)
                    elif 1 > abs(cell.value) < 1e-4:  # Small numbers
                        cell_length = len("%.5f" % cell.value)
                    else:
                        cell_length = len("%.3f" % cell.value)
                else:
                    cell_length = len(str(cell.value))
                if cell_length > max_length and cell.value is not None:
                    max_length = cell_length
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Align all cells. First column left aligned, meant for the index
    align_options = {'first_left': 'center', 'center': 'center', 'left': 'left'}
    if align not in align_options:
        raise ValueError("align must be 'first_left', 'center', or 'left'.")

    for column in worksheet.columns:
        current_align = align_options[align] if not (
                align == 'first_left' and column[0].column_letter == 'A') else 'left'
        for cell in column:
            if cell.row == 1 and header == "large" or wrap_column:
                wrap_text = True
            else:
                wrap_text = False
            cell.alignment = Alignment(horizontal=current_align, vertical='center', wrap_text=wrap_text)

    def is_column_empty(col):
        return all(c.value is None for c in col)

    # Add filters
    if filter == "all":  # This includes empty columns in the middle
        first_col = None
        last_col = None
        for column in worksheet.columns:
            if is_column_empty(column):
                continue
            if first_col is None:
                first_col = column[0].column
            last_col = column[-1].column

    elif filter == "first":
        for column in worksheet.columns:
            if not is_column_empty(column):
                first = column[0].column
                break
        else:
            raise ValueError("There are no non-empty columns.")
        first_col = last_col = first
    elif filter == "first_two":
        first_col = None
        last_col = None
        for column in worksheet.columns:
            if is_column_empty(column):
                continue
            if first_col is None:
                first_col = column[0].column
            elif last_col is None:
                last_col = column[0].column
                break
    elif filter == "last":
        last = max(column[0].column for column in worksheet.columns if not is_column_empty(column))
        first_col = last_col = last
    elif filter == "none":
        first_col = last_col = None
    else:
        raise ValueError("filter must be 'all', 'first', 'last', or 'none'.")

    if first_col is not None and last_col is not None:
        worksheet.auto_filter.ref = (f"{worksheet.cell(row=1, column=first_col).coordinate}:"
                                     f"{worksheet.cell(row=worksheet.max_row, column=last_col).coordinate}")

    # If filter is added, we need to adjust the width of the headers in case the width of the column = len(header).
    # Hardcoded values for the width of the headers
    if align == 'first_left' or align == "center":
        add_width = 5
    else:
        add_width = 3
    if filter != "none" and header != "exclude" and header != "large":
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            column_index = worksheet[column_letter + '1'].column
            if first_col <= column_index <= last_col:
                header_value = column[0].value
                if worksheet.column_dimensions[column_letter].width < len(str(header_value)) + add_width:
                    worksheet.column_dimensions[column_letter].width = len(str(header_value)) + add_width

    workbook.save(excel_path)
    # Open the Excel file
    if open_excel:
        os.system(f'start excel "{excel_path}"')


def generate_latex_table(df, output_path=None, precision=2, mathpazo=False, output="tex", caption=None, label=None):
    """Generate :math:`\\LaTeX` for a given DataFrame. Depending on the `output` parameter, it returns the :math:`\\LaTeX`
    code as a string, saves it to a .tex file, or compiles it to a PDF. In the latter case, it uses the `pdflatex` and
    `booktabs` packages for better aesthetics, so the output PDF is cropped to the table size.

    :param df: DataFrame containing the statistics.
    :type df: pd.DataFrame
    :param output_path: Path to the output if `output` is 'pdf' or 'tex', optional (default is `None`). It must be
        provided without an extension.
    :type output_path: str
    :param precision: Number of decimal places to show in the :math:`\\LaTeX` table, optional (default is 2).
    :type precision: int
    :param mathpazo: Whether to use the mathpazo font in the :math:`\\LaTeX` table, optional (default is `True`).
    :type mathpazo: bool
    :param output: Whether to output the :math:`\\LaTeX` code to a .tex file, a PDF file, or as a string, optional
        (default is `tex`). It can be `tex`, `pdf`, or `string`.
    :type output: str
    :param caption: Caption for the table, optional (default is `None`).
    :type caption: str
    :return: None or str
    :rtype: None or str


    .. note:: The `output_path` must be provided without an extension. The function will add the extension depending on
        the output parameter.

    """  # noqa: D205
    float_cols = df.select_dtypes(include=['float']).columns
    # Apply formatting only to numeric columns
    fmt = {col: f"{{:.{precision}f}}" for col in float_cols}
    column_format = '@{}' + 'c' * (len(df.columns)) + '@{}'
    latex_table = (
        df.style
        .hide(axis="index")
        .format(fmt, escape="latex")
        .to_latex(column_format=column_format, hrules=True)
    )
    if caption is not None:
        latex_table = (latex_table.replace(r'\begin{tabular}', r'\begin{table}[h]' + '\n' +
                                           r'\begin{tabular}') + '\n' + r'\caption{' + caption + '}'
                       + '\n' + r'\end{table}')
    # Add label at the end of the table
    if label is not None:
        latex_table = latex_table.replace(r'\end{table}', r'\label{' + label + '}' + '\n'
                                          + r'\end{table}')
    # Handle the output_path
    if output == "pdf" or output == "tex":
        end = output_path.split('.')[-1]
        split = output_path.split('.')
        msg = "The output file name contains dots. Make sure it is provided without an extension."
        if len(split) == 2:
            if end == output:  # No extension provided
                output_path += "." + output
            elif end == "tex" or end == "pdf":
                raise ValueError("The output file must be provided without an extension.")
            else:
                warnings.warn(msg)
                output_path = output_path + "." + output
        else:
            warnings.warn(msg)
            output_path = output_path + "." + output

    if output == "pdf":
        name = os.path.basename(output_path)
        path = os.path.dirname(output_path)
        os.chdir(path)
        with open('table.tex', 'w') as f:
            f.write(r'\documentclass{standalone}\usepackage{booktabs}')
            if mathpazo:
                f.write(r'\usepackage{mathpazo}')
            f.write(r'\begin{document}')
            f.write(latex_table)
            f.write(r'\end{document}')
        subprocess.run(['pdflatex', 'table.tex'])
        if os.path.exists(name):
            os.remove(name)
        os.rename('table.pdf', name)
        for ext in ['tex', 'log', 'aux']:  # Remove the auxiliary files
            os.remove(f'table.{ext}')

    elif output == "tex":
        with open(output_path, 'w') as f:
            f.write(latex_table)

    elif output == "string":
        return latex_table
    else:
        raise ValueError("output must be 'tex', 'pdf', or 'string'.")


def saving_routine_excel(wb, excel_path, open_excel=False):
    try:
        wb.save(excel_path)
    # Give the user the option to close the file if PermissionError
    except PermissionError:
        input(
            f"Permission denied: Unable to save the file at {excel_path}. Try closing the file and press Enter to continue...")
        wb.save(excel_path)
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        raise
    if open_excel:
        os.system(f'start excel "{excel_path}"')


# Particular of this validation, move to other script TODO

def rewrite_files(file_path, extension='.xlsx'):
    # Transform old names like this
    # (f'results_M{no_Vs}_randomV_perc_{percentage_random}_N_{no_random}_n_{n}_l_{l}_'
    #              f'addPhi_{additional_best_no}_search_{search}_less_{less_ability_coef}_noless_{no_less}'
    #              f'alternative_{alternative_stop}_stop_{stop}.xlsx')
    # to this new ones
    # (f'results_M_{no_Vs}_randomV_perc_{percentage_random}_N_{no_random}_n_{n}_l_{l}_'
    #              f'addPhi_{additional_best_no}_search_{search}_less_{less_ability_coef}_noless_{no_less}_'
    #              f'alternative_{alternative_stop}_stop_{stop}.xlsx')
    for file in os.listdir(file_path):
        if file.endswith(extension):
            # Extract the parameters from the file name
            split = file.split('_')
            if split[1] == 'M':
                continue
            new_file_name = f'results_M_{split[1][2:]}_randomVperc_{split[3]}_N_{split[5]}' + \
                            f'_n_{split[11]}_l_{split[13]}_addPhi_{split[7]}_search_{split[9]}_less_{split[15]}_' + \
                            f'noless_{split[17][0]}_alternative_{split[18]}_stop_{split[20]}'
            try:
                os.rename(os.path.join(file_path, file), os.path.join(file_path, new_file_name))
            except FileExistsError:
                print(f"File {new_file_name} already exists.")


def label_as_filename(file_path):
    """Given a path to a folder with .tex files, transform the caption into the filename."""
    for file in os.listdir(file_path):
        if file.endswith('.tex'):
            with open(os.path.join(file_path, file), 'r') as f:
                content = f.read()
                label = content.split('\\label{tab:')[1].split('}')[0]
                new_label = file.replace('.tex', '')
                content = content.replace(f'\\label{{tab:{label}}}', f'\\label{{tab:{new_label}}}')
                with open(os.path.join(file_path, file), 'w') as f:
                    f.write(content)


def filename_as_label(file_path):
    for file in os.listdir(file_path):
        if file.endswith('.tex'):
            with open(os.path.join(file_path, file), 'r') as f:
                content = f.read()
                label = content.split('\\label{tab:')[1].split('}')[0]
                new_name = label + '.tex'

            os.rename(os.path.join(file_path, file), os.path.join(file_path, new_name))


if __name__ == "__main__":
    path = os.path.join(os.path.dirname(__file__), 'results')
    # tables path is 3 levels up
    tables_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'tables')
    # rewrite_files(path, extension='.xlsx')
    # label_as_filename(tables_path)
